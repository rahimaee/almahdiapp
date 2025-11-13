from io import BytesIO
from django.http import HttpResponse
from urllib.parse import quote
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class ExcelExporter:
    def __init__(self, headers, data, required_fields=None, col_widths=None, rtl=True):
        self.headers = headers
        self.data = data
        self.required_fields = required_fields or []
        self.col_widths = col_widths or [15] * len(headers)
        self.rtl = rtl

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "نمونه واردسازی"

        # تنظیم جهت راست‌به‌چپ برای فایل‌های فارسی
        if rtl:
            try:
                self.ws.sheet_view.rightToLeft = True
            except Exception:
                pass

        # استایل‌ها
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        self.default_font = Font(bold=False)
        self.default_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="FFF3F3F3", end_color="FFF3F3F3", fill_type="solid")
        self.required_header_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        self.required_cell_fill = PatternFill(start_color="FFFDF2F2", end_color="FFFDF2F2", fill_type="solid")
        self.required_font = Font(bold=True, color="B30000")
        
        thin = Side(border_style="thin", color="000000")
        self.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # محاسبه اندیس ستون‌های الزامی
        self.required_idxs = [i for i, h in enumerate(headers) if h in self.required_fields]

    def _set_cell(self, row, col, value, is_header=False):
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.alignment = self.center_align
        cell.border = self.border  # اضافه کردن حاشیه به همه سلول‌ها

        if is_header:
            if (col - 1) in self.required_idxs:
                cell.fill = self.required_header_fill
                cell.font = self.required_font
            else:
                cell.fill = self.header_fill
                cell.font = self.header_font
        else:
            cell.fill = self.default_fill
            cell.font = self.default_font
            if (col - 1) in self.required_idxs:
                cell.fill = self.required_cell_fill
                cell.font = self.required_font

    def export_to_bytes(self):
        # نوشتن header
        for col_idx, header in enumerate(self.headers, start=1):
            self._set_cell(row=1, col=col_idx, value=header, is_header=True)

        # نوشتن داده‌ها
        for row_idx, row_data in enumerate(self.data, start=2):
            if isinstance(row_data, dict):
                for col_idx, field in enumerate(self.headers, start=1):
                    value = row_data.get(field, "")
                    self._set_cell(row=row_idx, col=col_idx, value=value)
            elif isinstance(row_data, (list, tuple)):
                for col_idx, value in enumerate(row_data, start=1):
                    self._set_cell(row=row_idx, col=col_idx, value=value)

        # تنظیم خودکار عرض ستون‌ها بر اساس محتوای هر ستون
        for col_idx, col_cells in enumerate(self.ws.columns, start=1):
            max_length = 0
            for cell in col_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            # افزودن حاشیه 2 کاراکتری برای زیبایی
            self.ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        # تنظیم فونت پیش‌فرض کل شیت به Vazir
        self.default_font = Font(name='Vazir', bold=False)
        self.header_font = Font(name='Vazir', bold=True)
        self.required_font = Font(name='Vazir', bold=True, color="B30000")

        # ارتفاع سطرها
        self.ws.row_dimensions[1].height = 28
        for row_idx in range(2, len(self.data) + 2):
            self.ws.row_dimensions[row_idx].height = 22

        # ذخیره در بافر
        bio = BytesIO()
        self.wb.save(bio)
        bio.seek(0)
        return bio

    def response(bio: BytesIO, filename: str) -> HttpResponse:
        """
        ساخت پاسخ HTTP برای دانلود فایل Excel با نام فارسی.
        """
        quoted = quote(filename)
        response = HttpResponse(
            bio.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quoted}; filename=\"sample.xlsx\""
        return response

import pandas as pd
from io import BytesIO

class ExcelImport:
    def __init__(self, file, choices=None, required_fields=None):
        """
        file: instance از request.FILES
        choices: لیست tuple [('field_key', 'Label')] برای mapping ستون‌ها
        required_fields: ستون‌های الزامی بر اساس کلید
        """
        self.file = file
        self.choices = choices or []
        self.required_fields = required_fields or []
        self._df = None
        self._records = None

    def read_file(self):
        """
        خواندن فایل اکسل از UploadedFile و تبدیل همه ستون‌ها به رشته
        """
        # بافر از فایل آپلود شده
        bio = BytesIO(self.file.read())
        self._df = pd.read_excel(bio, sheet_name=0, dtype=str)
        self._df = self._df.fillna("")  # تبدیل نان به رشته خالی
        return self._df

    def clean_data(self):
        """
        trim کردن مقادیر متنی و حذف ردیف‌های خالی
        """
        if self._df is None:
            raise ValueError("ابتدا باید فایل خوانده شود.")
        
        # حذف ردیف‌های کاملاً خالی
        self._df.dropna(how="all", inplace=True)
        
        # trim همه مقادیر متنی
        for col in self._df.select_dtypes(include="object").columns:
            self._df[col] = self._df[col].str.strip()
        
        # تغییر نام ستون‌ها بر اساس choices
        if self.choices:
            choices_dict = dict(self.choices)
            swapped_choices_dict = {v: k for k, v in choices_dict.items()}
            self._df.columns = [swapped_choices_dict.get(col, col) for col in self._df.columns]

        return self._df

    @property
    def records(self):
        """
        لیست دیکشنری‌ها از داده خوانده شده
        لزوماً clean_data اجرا نمی‌شود
        """
        if self._records is None:
            if self._df is None:
                raise ValueError("ابتدا باید فایل خوانده شود.")
            self._records = self._df.to_dict(orient="records")
        return self._records
