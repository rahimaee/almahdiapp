import jdatetime
import traceback
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from accounts_apps.decorators import feature_required
from accounts_apps.utils import get_accessible_soldiers_for_user
from soldier_vacation_apps.models import LeaveBalance
from soldire_letter_apps.models import NormalLetterMentalHealthAssessmentAndEvaluation, IntroductionLetter, \
    NormalLetter, NormalLetterHealthIodine
from .models import Soldier, OrganizationalCode, Settlement, PaymentReceipt
from .forms import *
from cities_iran_manager_apps.models import City, Province
from units_apps.models import ParentUnit, SubUnit
import jdatetime
from soldier_documents_apps.models import SoldierDocuments
from soldier_service_apps.models import SoldierService
from apps_settings.models import AppsSettings
import zipfile
import os
from django.shortcuts import render
from django.core.files.base import ContentFile
from .models import Soldier
from .forms import PhotoZipUploadForm
from datetime import date
from django.utils.timezone import now
from .models import UnitHistory
from django.db.models.fields.related import ForeignKey, OneToOneField
from django.core.paginator import Paginator
import openpyxl
from django.shortcuts import render, redirect
from django.db import transaction
from django.http import HttpResponse
from .models import Soldier, OrganizationalCode
from .forms import SoldierUploadForm
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .utils import create_soldiers_excel
from datetime import datetime
import jdatetime
from .enums import SoldierStatusFilterEnum
def jalali_to_gregorian(jdate_str):
    y, m, d = map(int, jdate_str.split('/'))
    return jdatetime.date(y, m, d).togregorian()

# @feature_required('لیست سربازان')
def soldier_list(request):
    default_filter = SoldierStatusFilterEnum.PRESENT.key  # مقدار پیش‌فرض
    status_choices = [(s.key, s.label) for s in SoldierStatusFilterEnum]

    form = SoldierSearchForm(request.GET or None)
    soldiers = get_accessible_soldiers_for_user(request.user).order_by('organizational_code__code_number')
    soldiers = soldiers.filter().select_related(
        'residence_province',
        'residence_city',
        'current_parent_unit',
        'current_sub_unit',
        'basic_training_center'
    )
    # مقدار انتخاب شده توسط کاربر
    selected_filter = request.GET.get("defaultFilter", default_filter)
    complex_query = request.GET.get("complex_query", '')
    # ================================
    #   اعمال فیلتر وضعیت انتخاب شده
    # ================================
    if selected_filter == SoldierStatusFilterEnum.ALL.key:
        # بدون فیلتر
        pass
    elif selected_filter == SoldierStatusFilterEnum.PRESENT_AND_ABSENT.key:
        soldiers = soldiers.filter(is_checked_out=False)

    elif selected_filter == SoldierStatusFilterEnum.PRESENT.key:
        soldiers = soldiers.filter(is_fugitive=False, is_checked_out=False)

    elif selected_filter == SoldierStatusFilterEnum.ABSENT.key:
        soldiers = soldiers.filter(is_fugitive=True, is_checked_out=False)

    elif selected_filter == SoldierStatusFilterEnum.CHECKOUT.key:
        soldiers = soldiers.filter(is_checked_out=True)
    selected_filter_label = next(
        (item.label for item in SoldierStatusFilterEnum if item.key == selected_filter),
        ""
    )
    all_soldiers_counts= len(soldiers) 
    remaining_filter = None
    
    if form.is_valid():
        print("isvald  form")
        data = form.cleaned_data

        # ساخت کوئری پویا
        query = Q()

        # فیلدهای متنی
        text_fields = [
            'national_code', 'first_name', 'last_name', 'father_name',
            'id_card_code', 'birth_place',
            'issuance_place', 'health_status_description'
        ]
        for field in text_fields:
            if value := data.get(field):
                query &= Q(**{f"{field}__icontains": value})

        # فیلدهای رابطه‌ای
        relation_fields = [
            'residence_province', 'residence_city',
            'current_parent_unit', 'current_sub_unit',
            'basic_training_center'
        ]
        for field in relation_fields:
            if value := data.get(field):
                query &= Q(**{field: value})

        # فیلدهای انتخابی
        choice_fields = [
            'rank', 'health_status', 'blood_group',
            'degree', 'skill_certificate', 'skill_group',
            'has_driving_license', 'marital_status'
        ]
        for field in choice_fields:
            if value := data.get(field):
                query &= Q(**{field: value})

        # فیلدهای عددی
        numeric_fields = [
            'training_duration', 'essential_service_duration',
            'number_of_children', 'number_of_certificates'
        ]
        for field in numeric_fields:
            if value := data.get(field):
                query &= Q(**{field: value})

        # فیلدهای تاریخی
        date_fields = ['service_entry_date', 'dispatch_date', 'birth_date']
        for field in date_fields:
            value = data.get(field)
            if value:
                try:
                    # تبدیل تاریخ شمسی به میلادی
                    year, month, day = map(int, value.split('/'))
                    gregorian_date = jdatetime.date(year, month, day).togregorian()
                    query &= Q(**{f"{field}__date": gregorian_date})
                except (ValueError, TypeError):
                    pass

        # فیلدهای بولی
        bool_fields = [
            'is_guard_duty', 'independent_married', 'is_certificate'
        ]
        for field in bool_fields:
            if value := data.get(field):
                query &= Q(**{field: value})

        # ---------------- فیلتر تاریخ پایان خدمت ----------------
        end_from = data.get("end_service_from_date")
        end_to   = data.get("end_service_to_date")

        if end_from:
            query &= Q(service_end_date__gte=end_from)
        if end_to:
            query &= Q(service_end_date__lte=end_to)
        # ---------------- فیلتر روز باقی‌مانده تا پایان خدمت ----------------

        if remaining_filter == "unknown":
            query &= Q(remaining_days__isnull=True)

        elif remaining_filter == "end":
            query &= Q(remaining_days=0)

        elif remaining_filter == "remaining":
            query &= Q(remaining_days__gt=45)

        elif remaining_filter == "remaining45":
            query &= Q(remaining_days__lte=45)

        elif remaining_filter == "remaining30":
            query &= Q(remaining_days__lte=30)

        elif remaining_filter == "remaining15":
            query &= Q(remaining_days__lte=15)

        elif remaining_filter == "remaining5":
            query &= Q(remaining_days__lte=5)

        organizational_code = request.GET.get("organizational_code", '')
        if organizational_code:
            query &= Q(organizational_code__code_number=organizational_code)
        # اعمال فیلتر
        soldiers = soldiers.filter(query)


    if complex_query:
        query |= Q(organizational_code__code_number__icontains=complex_query)
        query |= Q(national_code__icontains=complex_query)
        query |= Q(first_name__icontains=complex_query)
        query |= Q(last_name__icontains=complex_query)

        # Full name search
        parts = complex_query.split()

        if len(parts) == 2:
            first, last = parts
            query |= Q(first_name__icontains=first, last_name__icontains=last)

        soldiers = soldiers.filter(query)

    action = request.POST.get("action")

    if action == "exportexcel":
        wb = create_soldiers_excel(soldiers)

        response = HttpResponse(
            content=openpyxl.writer.excel.save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="soldiers.xlsx"'
        return response    
    # ---------------- Pagination ----------------
    page_number = request.GET.get('page', 1)
    paginator = Paginator(soldiers, 20)  # 20 سرباز در هر صفحه
    page_obj = paginator.get_page(page_number)

    return render(request, 'soldires_apps/soldier_list.html', {
        'form': form,
        'soldiers': page_obj,  # لیست سربازان برای صفحه جاری
        'paginator': paginator,
        'page_obj': page_obj,
        'soldiers_counts' : len(soldiers),
        'all_soldiers_counts':all_soldiers_counts,
        'remainingFilter':remaining_filter,
        'status_choices':status_choices,
        'selected_filter':selected_filter,
        'selected_filter_label':selected_filter_label,
        'complex_query':complex_query or '',
    })


def soldiers_date_to_end(request):
    """
    نمایش سربازانی که تا پایان خدمتشان روز مشخصی مانده
    """
    days_filter = request.GET.get("days")  # مقدار روز از کوئری استرینگ

    soldiers = Soldier.objects.filter(is_checked_out=False)

    if days_filter:
        try:
            days = int(days_filter)
            soldiers = Soldier.date_to_ends(days)  # استفاده از متد کلاس
        except ValueError:
            pass  # اگر مقدار غیر عددی بود نادیده گرفته شود

    return render(request, "soldires_apps/soldiers_date_to_ends.html", {
        "soldiers": soldiers,
        "days_filter": days_filter or "",
    })

def convert_jalali_to_gregorian_string(jalali_str):
    """تبدیل رشته تاریخ شمسی به رشته میلادی به فرمت YYYY-MM-DD"""
    try:
        parts = [int(part) for part in jalali_str.split('/')]
        j_date = jdatetime.date(*parts)
        g_date = j_date.togregorian()
        return g_date.strftime('%Y-%m-%d')
    except Exception as e:
        return None  # یا raise e


def soldier_create(request):
    if request.method == "POST":
        post_data = request.POST.copy()

        # تبدیل تاریخ‌های شمسی به میلادی
        date_fields = ["birth_date", "dispatch_date", "service_entry_date"]
        for field in date_fields:
            date_value = post_data.get(field)
            if date_value:
                try:
                    year, month, day = map(int, date_value.split("/"))
                    gregorian_date = jdatetime.date(year, month, day).togregorian()
                    post_data[field] = gregorian_date.isoformat()
                except ValueError:
                    pass

        # ساخت فرم
        form = SoldierForm(post_data)

        # بروزرسانی شهرها و زیرواحدها در فرم
        try:
            province = Province.objects.get(id=post_data.get('residence_province'))
            form.fields['residence_city'].queryset = City.objects.filter(province=province)
        except (Province.DoesNotExist, ValueError, TypeError):
            form.fields['residence_city'].queryset = City.objects.none()

        try:
            parent_unit = ParentUnit.objects.get(id=post_data.get('current_parent_unit'))
            form.fields['current_sub_unit'].queryset = SubUnit.objects.filter(parent_unit=parent_unit)
        except (ParentUnit.DoesNotExist, ValueError, TypeError):
            form.fields['current_sub_unit'].queryset = SubUnit.objects.none()

        if form.is_valid():
            try:
                with transaction.atomic():
                    soldier = form.save()
                    aps = AppsSettings.objects.first()

                    # تخصیص کد سازمانی
                    org_code = soldier.organizational_code
                    if org_code.is_active:
                        new_code = OrganizationalCode.objects.filter(is_active=False).first()
                        if not new_code:
                            form.add_error(None, "کد آزاد برای اختصاص به سرباز موجود نیست.")
                            raise IntegrityError("No free organizational code available.")
                        soldier.organizational_code = new_code
                        soldier.save()
                        org_code = new_code
                    org_code.is_active = True
                    org_code.save()

                    # مدارک و سرویس
                    SoldierDocuments.objects.create(soldier=soldier)

                    # ایجاد سرویس سرباز با روش جدید
                    try:
                        service = SoldierService()
                        service.soldier = soldier
                        service.save()

                        # کسری بر اساس وضعیت تأهل
                        if soldier.marital_status == 'متاهل':
                            service.reduction_spouse = 60
                            children = soldier.number_of_children
                            if children >= 1:
                                service.reduction_children = 90
                            if children >= 2:
                                service.reduction_children += 120
                            if children >= 3:
                                service.reduction_children += 150
                            service.save()
                    except Exception as e:
                        print(f"Error creating SoldierService: {str(e)}")
                        raise

                    # ایجاد نامه‌های مربوطه
                    IntroductionLetter.objects.create(
                        soldier=soldier,
                        part=form.cleaned_data['current_parent_unit'],
                        sub_part=form.cleaned_data['current_sub_unit'],
                        letter_type='چهاربرگ+معرفی نامه'
                    )

                    normal_letter = NormalLetter.objects.create(
                        soldier=soldier,
                        destination='به : قسمت بهداشت و درمان آموزشگاه رزم مقدماتی المهدی (عج) نیروی زمینی سپاه',
                        letter_type='سنجش و ارزیابی سلامت روان',
                        created_by=request.user if request.user.is_authenticated else None
                    )

                    NormalLetterMentalHealthAssessmentAndEvaluation.objects.create(
                        subject='تست سلامت بدو ورود',
                        normal_letter=normal_letter
                    )
                    if soldier.current_sub_unit and getattr(soldier.current_sub_unit, 'HealthIodine', False):
                        letter2 = NormalLetter.objects.create(
                            soldier=soldier,
                            destination='به : قسمت بهداشت و درمان آموزشگاه رزم مقدماتی المهدی (عج) نیروی زمینی سپاه',
                            letter_type='دریافت تائیدیه سلامت',
                            created_by=request.user if request.user.is_authenticated else None
                        )
                        NormalLetterHealthIodine.objects.create(
                            normal_letter=letter2,
                            sub_part=soldier.current_sub_unit,
                            part=soldier.current_parent_unit
                        )
                    return redirect("soldier_list")
            except IntegrityError as e:
                print("IntegrityError occurred:", e)
                traceback.print_exc()
                form.add_error(None, f"خطای پایگاه داده: {str(e)}")
            except Exception as e:
                print("Error occurred:", e)
                traceback.print_exc()
                form.add_error(None, f"خطای غیرمنتظره: {str(e)}")

    else:
        form = SoldierForm()
        aps = AppsSettings.objects.first()
        form.fields['essential_service_duration'].initial = aps.essential_service_duration if aps else 0

    return render(request, "soldires_apps/soldier_create.html", {"form": form})


def soldier_detail(request, pk):
    # دریافت سرباز بر اساس primary key (pk)
    soldier = get_object_or_404(Soldier, pk=pk)
    return render(request, 'soldires_apps/soldier_detail.html', {'soldier': soldier})


def soldier_edit(request, pk):
    soldier = get_object_or_404(Soldier, pk=pk)

    if request.method == "POST":
        post_data = request.POST.copy()

        # تبدیل تاریخ‌ها
        date_fields = ["birth_date", "dispatch_date", "service_entry_date"]
        for field in date_fields:
            date_value = post_data.get(field)
            if date_value:
                try:
                    date_value = date_value.strip()
                    if '/' in date_value:
                        year, month, day = map(int, date_value.split("/"))
                        gregorian_date = jdatetime.date(year, month, day).togregorian()
                    else:
                        year, month, day = map(int, date_value.split("-"))
                        gregorian_date = date(year, month, day)
                    post_data[field] = gregorian_date
                except:
                    post_data[field] = getattr(soldier, field)

        # نگهداری مقادیر قبلی واحدها
        old_parent_unit = soldier.current_parent_unit
        old_sub_unit = soldier.current_sub_unit

        form = SoldierFormUpdate(post_data, instance=soldier)

        # وابستگی‌ها
        if post_data.get('residence_province'):
            try:
                province = Province.objects.get(id=post_data['residence_province'])
                form.fields['residence_city'].queryset = City.objects.filter(province=province)
            except:
                form.fields['residence_city'].queryset = City.objects.none()

        if post_data.get('current_parent_unit'):
            try:
                parent_unit = ParentUnit.objects.get(id=post_data['current_parent_unit'])
                form.fields['current_sub_unit'].queryset = SubUnit.objects.filter(parent_unit=parent_unit)
            except:
                form.fields['current_sub_unit'].queryset = SubUnit.objects.none()

        form.fields['organizational_code'].queryset = OrganizationalCode.objects.filter(
            Q(is_active=False) | Q(id=soldier.organizational_code.id)
        )

        if form.is_valid():
            updated_fields = {}
            for field_name in form.fields:
                updated_fields[field_name] = form.cleaned_data.get(field_name)

            for field, value in updated_fields.items():
                setattr(soldier, field, value)

            soldier.save()

            # ثبت تاریخچه اگر تغییر کرده باشد
            if old_parent_unit != soldier.current_parent_unit or old_sub_unit != soldier.current_sub_unit:
                last_history = soldier.unit_history.filter(end_date__isnull=True).last()
                if last_history:
                    last_history.end_date = now()
                    last_history.save()

                UnitHistory.objects.create(
                    soldier=soldier,
                    parent_unit=soldier.current_parent_unit,
                    sub_unit=soldier.current_sub_unit,
                    start_date=now(),
                )

            return redirect('soldier_detail', pk=soldier.pk)
        else:
            print("Form errors:", form.errors)

    else:
        form = SoldierFormUpdate(instance=soldier)

        if soldier.residence_province:
            form.fields['residence_city'].queryset = City.objects.filter(province=soldier.residence_province)
        if soldier.current_parent_unit:
            form.fields['current_sub_unit'].queryset = SubUnit.objects.filter(parent_unit=soldier.current_parent_unit)

        form.fields['organizational_code'].queryset = OrganizationalCode.objects.filter(
            Q(is_active=False) | Q(id=soldier.organizational_code.id)
        )

        # تاریخ‌های شمسی
        if soldier.birth_date:
            j = jdatetime.date.fromgregorian(date=soldier.birth_date)
            form.initial['birth_date'] = f"{j.year}/{str(j.month).zfill(2)}/{str(j.day).zfill(2)}"
        if soldier.dispatch_date:
            j = jdatetime.date.fromgregorian(date=soldier.dispatch_date)
            form.initial['dispatch_date'] = f"{j.year}/{str(j.month).zfill(2)}/{str(j.day).zfill(2)}"
        if soldier.service_entry_date:
            j = jdatetime.date.fromgregorian(date=soldier.service_entry_date)
            form.initial['service_entry_date'] = f"{j.year}/{str(j.month).zfill(2)}/{str(j.day).zfill(2)}"

    return render(request, 'soldires_apps/soldier_edit.html', {'form': form})


def upload_soldier_photo(request, soldier_id):
    soldier = get_object_or_404(Soldier, id=soldier_id)

    if request.method == 'POST':
        form = SoldierPhotoForm(request.POST, request.FILES, instance=soldier)
        if form.is_valid():
            form.save()
            return redirect('soldier_detail', pk=soldier.id)  # یا هر صفحه‌ای که می‌خواهی برگردد
    else:
        form = SoldierPhotoForm(instance=soldier)

    return render(request, 'soldires_apps/upload_photo.html', {'form': form, 'soldier': soldier})


def bulk_photo_upload(request):
    message = ""
    not_found = []  # کدملی‌هایی که پیدا نشدند
    uploaded = []  # کدملی‌هایی که موفق بودن

    if request.method == 'POST':
        form = PhotoZipUploadForm(request.POST, request.FILES)
        if form.is_valid():
            zip_file = request.FILES['zip_file']
            if zipfile.is_zipfile(zip_file):
                zip_data = zipfile.ZipFile(zip_file)
                for filename in zip_data.namelist():
                    if filename.endswith(('.jpg', '.jpeg', '.png')):
                        national_code = os.path.splitext(os.path.basename(filename))[0].strip()
                        try:
                            soldier = Soldier.objects.get(national_code=national_code)
                            image_data = zip_data.read(filename)
                            soldier.photo_scan.save(
                                filename,
                                ContentFile(image_data),
                                save=True
                            )
                            uploaded.append(national_code)
                        except Soldier.DoesNotExist:
                            not_found.append(national_code)
                message = "عملیات آپلود تمام شد."
            else:
                message = "فایل ZIP معتبر نیست."
    else:
        form = PhotoZipUploadForm()

    return render(request, 'soldires_apps/bulk_photo_upload.html', {
        'form': form,
        'message': message,
        'uploaded': uploaded,
        'not_found': not_found
    })


def review_settlements(request):
    # فقط settlementهایی که در وضعیت بررسی حقوق هستند نمایش داده شوند
    settlements = Settlement.objects.filter(status='review').select_related('soldier')

    if request.method == 'POST':
        for settlement in settlements:
            new_debt = request.POST.get(f'debt_{settlement.id}')
            no_review = request.POST.get(f'no_review_{settlement.id}')
            approve_settlement = request.POST.get(f'approve_{settlement.id}')

            if new_debt and new_debt.strip():
                try:
                    debt_amount = int(new_debt)
                    settlement.total_debt_rial = debt_amount
                    settlement.current_debt_rial = debt_amount
                    settlement.last_rights_check_date = timezone.now().date()
                    settlement.updated_at = timezone.now()
                    # اگر بدهی صفر باشد، وضعیت تسویه کامل شود
                    if debt_amount == 0:
                        settlement.status = 'cleared'
                        settlement.need_rights_recheck = False
                    # اگر تیک نیاز به بررسی ندارد خورده باشد، وضعیت به pending تغییر کند
                    elif no_review:
                        settlement.status = 'pending'
                        settlement.need_rights_recheck = False
                    # در غیر این صورت، وضعیت همچنان review بماند
                    else:
                        settlement.status = 'review'
                        settlement.need_rights_recheck = True
                    settlement.save()
                    messages.success(request, f"بدهی سرباز {settlement.soldier.first_name} {settlement.soldier.last_name} به {debt_amount:,} ریال تنظیم شد.")
                except ValueError:
                    messages.error(request, f"مقدار بدهی برای سرباز {settlement.soldier.first_name} {settlement.soldier.last_name} نامعتبر است.")
            elif no_review:
                settlement.need_rights_recheck = False
                settlement.status = 'pending'  # تغییر وضعیت به در انتظار تسویه
                settlement.updated_at = timezone.now()
                settlement.save()
                messages.info(request, f"بررسی سرباز {settlement.soldier.first_name} {settlement.soldier.last_name} لغو شد و به مرحله تسویه منتقل شد.")
            elif approve_settlement:
                settlement.need_rights_recheck = False
                settlement.status = 'pending'
                settlement.updated_at = timezone.now()
                settlement.save()
                messages.success(request, f"تسویه‌حساب سرباز {settlement.soldier.first_name} {settlement.soldier.last_name} تایید شد.")
        return redirect('review_settlements')
    return render(request, 'soldires_apps/review_settlements.html', {'settlements': settlements})


def payment_receipt_create(request):
    # گرفتن settlement_id از URL اگر وجود داشته باشد
    settlement_id = request.GET.get('settlement_id')
    if settlement_id:
        # اگر settlement_id مشخص شده، فقط آن settlement را نمایش بده
        try:
            selected_settlement = Settlement.objects.get(id=settlement_id, current_debt_rial__gt=0)
            settlements = [selected_settlement]
        except Settlement.DoesNotExist:
            print(Settlement.DoesNotExist)
            messages.error(request, "تسویه‌حساب مورد نظر یافت نشد یا بدهی ندارد.")
            return redirect('soldiers_settlement_list')
    else:
        # در غیر این صورت، همه settlementهایی که بدهی دارند را نمایش بده
        settlements = Settlement.objects.filter(current_debt_rial__gt=0).select_related('soldier')

    if request.method == 'POST':
        # گرفتن اطلاعات از فرم
        settlement_id = request.POST.get('settlement_id')
        amount_rial = request.POST.get('amount_rial')
        receipt_number = request.POST.get('receipt_number')
        deposit_date = request.POST.get('deposit_date')
        bank_operator_code = request.POST.get('bank_operator_code')

        try:
            # اعتبارسنجی داده‌ها
            if not all([settlement_id, amount_rial, receipt_number, deposit_date, bank_operator_code]):
                messages.error(request, "لطفاً تمام فیلدها را پر کنید.")
                return render(request, 'soldires_apps/payment_receipt_create.html', {'settlements': settlements})

            amount_rial = int(amount_rial)
            if amount_rial <= 0:
                messages.error(request, "مبلغ واریزی باید بیشتر از صفر باشد.")
                return render(request, 'soldires_apps/payment_receipt_create.html', {'settlements': settlements})

            # پیدا کردن تسویه‌حساب سرباز انتخاب شده
            settlement = Settlement.objects.get(id=settlement_id)
            
            # بررسی اینکه مبلغ واریزی از بدهی باقی‌مانده بیشتر نباشد
            if amount_rial > settlement.current_debt_rial:
                messages.warning(request, f"مبلغ واریزی ({amount_rial:,} ریال) از بدهی باقی‌مانده ({settlement.current_debt_rial:,} ریال) بیشتر است.")

            # ایجاد فیش واریزی
            PaymentReceipt.objects.create(
                settlement=settlement,
                amount_rial=amount_rial,
                receipt_number=receipt_number,
                deposit_date=deposit_date,
                bank_operator_code=bank_operator_code
            )
            
            messages.success(request, f"فیش واریزی {receipt_number} به مبلغ {amount_rial:,} ریال برای سرباز {settlement.soldier.first_name} {settlement.soldier.last_name} با موفقیت ثبت شد.")
            
            # هدایت به صفحه مشاهده فیش‌های سرباز
            return redirect('settlement_payments', settlement_id=settlement.id)
            
        except ValueError:
            messages.error(request, "مبلغ واریزی باید عدد صحیح باشد.")
        except Settlement.DoesNotExist:
            messages.error(request, "تسویه‌حساب مورد نظر یافت نشد.")
        except Exception as e:
            messages.error(request, f"خطا در ثبت فیش واریزی: {str(e)}")

    return render(request, 'soldires_apps/payment_receipt_create.html', {
        'settlements': settlements,
        'selected_settlement_id': settlement_id,
        'today_date': timezone.now().date()
    })


def soldiers_settlement_list(request):
    # گرفتن فقط سربازانی که Settlement دارند و وضعیتشان review نیست
    settlements = Settlement.objects.select_related('soldier').exclude(status='review').order_by('-created_at')
    
    # فیلتر بر اساس وضعیت
    status_filter = request.GET.get('status')
    if status_filter:
        settlements = settlements.filter(status=status_filter)
    
    # فیلتر بر اساس بدهی
    debt_filter = request.GET.get('debt')
    if debt_filter == 'has_debt':
        settlements = settlements.filter(current_debt_rial__gt=0)
    elif debt_filter == 'no_debt':
        settlements = settlements.filter(current_debt_rial=0)

    # گرفتن اطلاعات تسویه‌حساب برای هر سرباز
    soldier_data = []
    for settlement in settlements:
        soldier_data.append({
            'soldier': settlement.soldier,
            'settlement': settlement
        })

    return render(request, 'soldires_apps/soldiers_settlement_list.html', {
        'soldier_data': soldier_data,
        'status_filter': status_filter,
        'debt_filter': debt_filter
    })


def settlement_payments_view(request, settlement_id):
    settlement = get_object_or_404(Settlement, id=settlement_id)
    payments = settlement.payments.all().order_by('-deposit_date')  # آخرین‌ها اول
    return render(request, 'soldires_apps/settlement_payments.html', {
        'settlement': settlement,
        'payments': payments,
    })


from django.db.models import OuterRef, Subquery, Max


def due_mental_health_letters(request):
    six_months_ago = timezone.now() - timedelta(days=180)

    # آخرین تست هر سرباز
    latest_letters = NormalLetterMentalHealthAssessmentAndEvaluation.objects.values(
        'normal_letter__soldier'
    ).annotate(
        latest_date=Max('created_at')
    ).filter(
        latest_date__lte=six_months_ago,
        normal_letter__soldier__is_checked_out=False,
    )

    # گرفتن آی‌دی سربازهایی که شرایط بالا رو دارند
    due_soldier_ids = [entry['normal_letter__soldier'] for entry in latest_letters]

    # پیدا کردن آخرین نامه‌های آن سربازها
    due_letters = NormalLetterMentalHealthAssessmentAndEvaluation.objects.filter(
        normal_letter__soldier__in=due_soldier_ids
    ).order_by('normal_letter__soldier', '-created_at').distinct('normal_letter__soldier')

    context = {
        'due_letters': due_letters,
    }
    return render(request, 'soldires_apps/due_mental_health_letters.html', context)


def soldiers_by_entry_date(request):
    soldiers = []
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        try:
            # تبدیل تاریخ‌های شمسی به میلادی
            from_parts = list(map(int, from_date.split('/')))
            to_parts = list(map(int, to_date.split('/')))
            from_gregorian = jdatetime.date(*from_parts).togregorian()
            to_gregorian = jdatetime.date(*to_parts).togregorian()

            soldiers = Soldier.objects.filter(
                service_entry_date__range=(from_gregorian, to_gregorian),
                # is_checked_out=False,
            ).order_by('service_entry_date')

        except Exception as e:
            print("خطا در تبدیل تاریخ:", e)

    return render(request, 'soldires_apps/soldiers_by_entry_date.html', {
        'soldiers': soldiers,
        'from_date': from_date,
        'to_date': to_date
    })


def incomplete_soldiers_list(request):
    from .models import Soldier
    soldiers = Soldier.objects.filter(is_checked_out=False).all()
    return render(request, 'soldires_apps/incomplete_soldiers_list.html', {'soldiers': soldiers})


def checked_out_soldiers_list(request):
    """لیست سربازانی که تسویه‌حساب شده‌اند"""
    from .models import Soldier
    soldiers = Soldier.objects.filter(is_checked_out=True).select_related(
        'residence_province',
        'residence_city',
        'current_parent_unit',
        'current_sub_unit'
    ).order_by('-status')
    
    return render(request, 'soldires_apps/checked_out_soldiers_list.html', {
        'soldiers': soldiers,
        'title': 'لیست سربازان تسویه‌حساب شده'
    })


def soldires_new_status_view(request, pk):
    soldier = get_object_or_404(Soldier, pk=pk)

    status_list_normal = ['ایجاد شده', 'چاپ و بررسی شده', 'تایید شده']
    status_list_intro = ['ایجاد شده', 'چاپ و درحال بررسی', 'تأیید نهایی']

    normal_letters = NormalLetter.objects.filter(soldier=soldier, status__in=status_list_normal)
    intro_letters = IntroductionLetter.objects.filter(soldier=soldier, status__in=status_list_intro)

    context = {
        'soldier': soldier,
        'normal_letters': normal_letters,
        'intro_letters': intro_letters,
        'status_list_normal': status_list_normal,
        'status_list_intro': status_list_intro,
    }

    return render(request, 'soldires_apps/soldires_new_status_view.html', context)

def organizational_codes_list(request):
    codes = OrganizationalCode.objects.annotate(
        soldiers_count=Count('soldiers')
    ).order_by('code_number')

    if request.method == "POST":
        form = OrganizationalCodeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "کد سازمانی با موفقیت اضافه شد ✅")
            return redirect('organizational_codes_list')
    else:
        form = OrganizationalCodeForm()

    return render(request, "soldires_apps/organizational_codes_list.html", {
        "codes": codes,
        "form": form,
    })
  
from django.shortcuts import render, redirect
from django.db import transaction
import openpyxl
from .models import Soldier, OrganizationalCode


def download_soldiers_template(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "سربازان"
    ws.sheet_view.rightToLeft = True

    headers = [
        "ردیف", "کد سازمانی", "کد ملی", "نام", "نام خانوادگی", "نام پدر",
        "تاریخ تولد", "وضعیت تاهل", "درجه", "وضعیت سلامت", 
        "تحصیلات", "آدرس", "موبایل",
        "شیعه", "قسمت", "زیر قسمت", "وضعیت تردد"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="soldiers_template.xlsx"'
    wb.save(response)
    return response

def is_jalali_leap(year):
    a = year - 474
    b = a % 2820 + 474
    return ((b + 38) * 682) % 2816 < 682

def jalali_to_gregorian(jalali_date_str):
    if not jalali_date_str:
        return None

    jalali_date_str = jalali_date_str.strip().replace('“', '').replace('”', '')
    parts = jalali_date_str.split('/')
    if len(parts) != 3:
        return None

    year, month, day = parts
    try:
        year = int(year)
        if year < 100:  # سال کوتاه 83 -> 1383
            year += 1300
        month = int(month)
        day = int(day)
    except ValueError:
        return None

    # محدود کردن روز به آخرین روز ماه شمسی
    if month <= 6:
        last_day = 31
    elif month <= 11:
        last_day = 30
    else:  # ماه 12
        last_day = 30 if is_jalali_leap(year) else 29
    day = min(day, last_day)

    # تبدیل به میلادی
    try:
        g_date = jdatetime.date(year, month, day).togregorian()
        return g_date.strftime("%Y-%m-%d")
    except ValueError:
        return None

import pandas as pd
from .utils import *

def soldiers_group_submit(request):
    form = SoldierUploadForm()
    soldiers_data = []

    if request.method == "POST":
        form = SoldierUploadForm(request.POST, request.FILES)
        action = request.POST.get('action')  # تشخیص نوع عملیات

        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file, sheet_name=0)

            # شروع از سطر دوم (index 1)
            for _, row in df.iloc[1:].iterrows():
                if not any(row):
                    continue

                soldier_info = {
                    'org_code': "" if pd.isna(row.get("کد سازمانی")) else row.get("کد سازمانی"),
                    'national_code': "" if pd.isna(row.get("کدملی")) else row.get("کدملی"),
                    'first_name': "" if pd.isna(row.get("نام")) else row.get("نام"),
                    'last_name': "" if pd.isna(row.get("نام خانوادگی")) else row.get("نام خانوادگی"),
                    'father_name': "" if pd.isna(row.get("نام پدر")) else row.get("نام پدر"),
                    'birth_date': "" if pd.isna(row.get("تاریخ تولد")) else row.get("تاریخ تولد"),
                    'marital_status': "" if pd.isna(row.get("وضعیت تاهل")) else row.get("وضعیت تاهل"),
                    'rank': "" if pd.isna(row.get("درجه")) else row.get("درجه"),
                    'health_status': "" if pd.isna(row.get("وضعیت سلامت")) else row.get("وضعیت سلامت"),
                    'degree': "" if pd.isna(row.get("مدرک")) else row.get("مدرک"),
                    'field_of_study': "" if pd.isna(row.get("رشته تحصیلی")) else row.get("رشته تحصیلی"),
                    'address': "" if pd.isna(row.get("آدرس منزل")) else row.get("آدرس منزل", row.get("محل سکونت", "")),
                    'phone_mobile': "" if pd.isna(row.get("موبایل")) else clean_phone(row.get("موبایل")),
                    'shiite': "" if pd.isna(row.get("شیعه")) else row.get("شیعه"),
                    'section': "" if pd.isna(row.get("نام قسمت")) else row.get("نام قسمت"),
                    'sub_section': "" if pd.isna(row.get("زیرقسمت")) else row.get("زیرقسمت"),
                    'traffic_status': "" if pd.isna(row.get("تردد")) else row.get("تردد"),
                }
                soldiers_data.append(soldier_info)

            # اگر مرحله بروزرسانی باشد، ذخیره داده‌ها
            if action == 'update':
                try:
                    import_soldiers_from_excel(excel_file)
                    messages.success(request, "اطلاعات سربازان با موفقیت ثبت شد.")
                except Exception as e:
                    print(e)
                    messages.error(request, f"خطا در ثبت اطلاعات: {str(e)}")

    context = {
        'form': form,
        'soldiers': soldiers_data,
    }
    return render(request, 'soldires_apps/soldiers_group_submit.html', context)

from django.db.models import Q
from django.http import JsonResponse

def soldiers_search(request):
    q = request.GET.get('q', '')
    soldiers = Soldier.objects.filter(
        Q(national_code__icontains=q) |
        Q(organizational_code__code_number__icontains=q) |
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q)
    )[:20]
    results = [
        {
            "id": s.id, 
            "text": f"کد ملی : {s.national_code} , {s.first_name} {s.last_name} [{s.organizational_code}]"
        }
        for s in soldiers
    ]
    return JsonResponse({"results": results})


from soldier_service_apps.models import SoldierService,get_service
from soldier_vacation_apps.models import LeaveBalance

def single_reports_soldier(request, soldier_id=None):
    soldier = get_object_or_404(Soldier, id=soldier_id)

    service = get_service(soldier)
    leave = get_object_or_404(LeaveBalance,soldier=soldier)

    # تاریخ امروز
    today = timezone.localdate()  # یا datetime.date.today()

    context = {
        'soldier': soldier,
        'service': service,
        'leave': leave,
        'today': today,
    }
    return render(request, 'soldires_apps/single_reports_soldier.html', context)


def import_organization_code_from_excel(request):
    
    pass

from django.shortcuts import redirect
from django.http import HttpResponse
from .models import OrganizationalCode, Soldier
from io import BytesIO
from almahdiapp.utils.excel import ExcelExporter,ExcelImport  # کلاس ExcelExporter که ارسال کردی
from almahdiapp.utils.builder import EnumMetaBuilder  # کلاس ExcelExporter که ارسال کردی
from django.contrib import messages
import pandas as pd
from .enums import OrganizationalCodeEnum,SoldierOrgCodeEnum 
from .enums import ExistingOrgCodeModeEnum,SoldierOrgCodeStatusEnum,OrganizationalCodeStatusEnum
from .constants import ORGANIZATIONAL_CODE_SAMPLE,SOLDIER_ORG_CODE_SAMPLE 

# =========================================================
# View اصلی فرم
# =========================================================
def organizational_code_match_view(request):
    """
    این view صفحه اصلی فرم را نشان می‌دهد و فرم‌ها را پردازش می‌کند.
    """
    status_choices = EnumMetaBuilder(OrganizationalCodeStatusEnum).choices
    existing_choices = EnumMetaBuilder(ExistingOrgCodeModeEnum).choices
    soldier_status_choices = EnumMetaBuilder(SoldierOrgCodeStatusEnum).choices

    status_default = OrganizationalCodeStatusEnum.INEXCEL
    existing_default = ExistingOrgCodeModeEnum.KEEP
    soldier_status_default = SoldierOrgCodeStatusEnum.INEXCEL
    
    context = {
        "status_choices":status_choices,
        "existing_choices":existing_choices,
        "soldier_status_choices":soldier_status_choices,
        "status_default":status_default,
        "existing_default":existing_default,
        "soldier_status_default":soldier_status_default,
    }
    
    return render(request, 'soldires_apps/organizational_codes_match.html', context)


# =========================================================
# دانلود فایل نمونه کد سازمانی
# =========================================================
def organizational_code_match_org_code_sample(request):
    eb = EnumMetaBuilder(OrganizationalCodeEnum)
    headers = eb.headers
    required_fields = [OrganizationalCodeEnum.NATIONAL_CODE.label, OrganizationalCodeEnum.ORG_CODE.label]

    data = [
        {e.label: row.get(e.key, "") for e in OrganizationalCodeEnum}
        for row in ORGANIZATIONAL_CODE_SAMPLE
    ]

    exporter = ExcelExporter(headers=headers, data=data, required_fields=required_fields)
    bio = exporter.export_to_bytes()
    return ExcelExporter.response(bio, "نمونه_کد_سازمانی.xlsx")


# =========================================================
# دانلود فایل نمونه تطبیق کد سازمانی با سرباز
# =========================================================
def organizational_code_match_soldier_org_code_sample(request):
    eb = EnumMetaBuilder(SoldierOrgCodeEnum)
    headers = eb.headers
    required_fields = [SoldierOrgCodeEnum.NATIONAL_CODE.label, SoldierOrgCodeEnum.ORG_CODE.label]

    data = [
        {e.label: row.get(e.key, "") for e in SoldierOrgCodeEnum}
        for row in SOLDIER_ORG_CODE_SAMPLE
    ]

    exporter = ExcelExporter(headers=headers, data=data, required_fields=required_fields)
    bio = exporter.export_to_bytes()
    return ExcelExporter.response(bio, "نمونه_تطبیق_سرباز.xlsx")

from django.db import transaction

def organizational_code_match_org_code(request):
    eb = EnumMetaBuilder(OrganizationalCodeEnum)
    
    if request.method != "POST":
        return redirect("organizational_codes_match")

    excel_file = request.FILES.get("excel_file")
    mode = request.POST.get("mode")
    existing_mode = request.POST.get("existing_mode")

    if not excel_file:
        messages.error(request, "هیچ فایلی انتخاب نشده است.")
        return redirect("organizational_codes_match")


    importer = ExcelImport(file=excel_file,choices=eb.choices)
    importer.read_file()
    importer.clean_data()
    records = importer.records

    processed_orgs = set()
    count = 0

    try:
        with transaction.atomic():  # تمام تغییرات در یک تراکنش
            for row in records:
                print(row)
                nc = row.get("national_code")
                oc = row.get("org_code")
                status = row.get("status")
                if not nc and not oc:
                    continue
                
                soldier = Soldier.objects.filter(national_code=nc).first()
                org = OrganizationalCode.objects.filter(code_number=oc).first()

                if not soldier and not org:
                    continue
                
                if not soldier and org:
                    soldier = org.current_soldier
                elif not org:
                    if soldier and soldier.organizational_code:
                        org = soldier.organizational_code
                    else:
                        continue

                print(org,"  ****  ",soldier)
                print(status,mode)    

                if mode == OrganizationalCodeStatusEnum.INEXCEL.key:
                    if status == OrganizationalCodeStatusEnum.ACTIVE.label:
                        org.current_soldier = soldier
                    elif status == OrganizationalCodeStatusEnum.INACTIVE.label:
                        org.current_soldier = None
                    elif status == OrganizationalCodeStatusEnum.CHECKOUT.label and soldier:
                        soldier.to_checkout()
                    elif status == OrganizationalCodeStatusEnum.PRESENT.key and soldier:
                        soldier.is_fugitive = False
                        soldier.is_checked_out = False
                        print("FF")
                    elif status == OrganizationalCodeStatusEnum.FUGITIVE.key and soldier:
                        soldier.is_fugitive = True
                        soldier.is_checked_out = False
                        print("TT")
                    soldier.save()
                        
                elif mode == OrganizationalCodeStatusEnum.ACTIVE.key:
                    org.current_soldier = soldier
                elif mode == OrganizationalCodeStatusEnum.INACTIVE.key:
                    org.current_soldier = None
                elif mode == OrganizationalCodeStatusEnum.CHECKOUT.key and soldier:
                    soldier.to_checkout()
                elif mode == OrganizationalCodeStatusEnum.PRESENT.key and soldier:
                        soldier.is_fugitive = False
                        soldier.is_checked_out = False
                        print("FF")
                elif mode == OrganizationalCodeStatusEnum.FUGITIVE.key and soldier:
                        soldier.is_fugitive = True
                        soldier.is_checked_out = False
                        print("TT")
                
                print("xXXX")    
                if soldier:
                    soldier.save()
                if org:
                    org.save()

                if org:
                    processed_orgs.add(org.id)
                count += 1

            remaining_orgs = OrganizationalCode.objects.exclude(id__in=processed_orgs)
            for org in remaining_orgs:
                soldier = org.current_soldier
                if existing_mode == ExistingOrgCodeModeEnum.ACTIVATE.key:
                    if not soldier:
                        org.current_soldier = Soldier.objects.filter(organizational_code=org).first()
                        org.save()
                elif existing_mode == ExistingOrgCodeModeEnum.DEACTIVATE.key:
                    if soldier:
                        org.current_soldier = None
                        org.save()
                # KEEP: بدون تغییر

    except Exception as e:
        print(e)
        messages.error(request, f"خطا در پردازش رکوردها: {str(e)}")
        return redirect("organizational_codes_match")

    messages.success(request, f"{count} رکورد با موفقیت پردازش شد.")
    return redirect("organizational_codes_match")

# ========================
# Form پردازش تطبیق کد سازمانی با سرباز
# ========================
def organizational_code_match_soldier_org_code(request):
    eb = EnumMetaBuilder(SoldierOrgCodeEnum)
    result = None

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        mode = request.POST.get("mode")

        if excel_file:
            importer = ExcelImport(file=excel_file,choices=eb.choices)
            importer.read_file()
            importer.clean_data()
            records = importer.records
            count = 0
            for row in records:
                nc = row.get("national_code")
                oc = row.get("org_code")
                new_status = row.get("new_status") or mode
                if not nc or not oc:
                    continue
                OrganizationalCode.objects.update_or_create(
                    national_code=nc,
                    defaults={"org_code": oc, "status": new_status}
                )
                count += 1

            messages.success(request, f"{count} رکورد با موفقیت تطبیق داده شد.")
        else:
            messages.error(request, "هیچ فایلی انتخاب نشده است.")

        return redirect("organizational_code_match_view")

    # GET: فقط رندر فرم
    return redirect("organizational_code_match_view")
