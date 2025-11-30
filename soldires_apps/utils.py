import pandas as pd
import jdatetime
from datetime import date
from django.db import transaction
from soldires_apps.models import Soldier, OrganizationalCode
from units_apps.models import ParentUnit, SubUnit
from accounts_apps.models import MyUser
from training_center_apps.models import TrainingCenter
from soldire_naserin_apps.models import NaserinGroup
from soldire_religious_period_apps.models import ReligiousPeriod
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from soldires_apps.models import Soldier
from django.db.models import Count, Q  # اضافه کردن Count
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

def clean_phone(value):
    if pd.isna(value) or value is None:
        return ""
    try:
        # اگر عدد بود، اعشار را حذف کن
        return str(int(float(value)))
    except (ValueError, TypeError):
        # اگر رشته است، فقط strip کن
        return str(value).strip()



def clean_int(value, default=0):
    if pd.isna(value):
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default



def shamsi_to_gregorian(jdate_str):
    if not jdate_str or pd.isna(jdate_str):
        return None
    try:
        parts = str(jdate_str).replace("-", "/").split("/")
        if len(parts) != 3:
            return None
        year, month, day = map(int, parts)
        if year < 100:
            year += 1300
        jdate = jdatetime.date(year, month, day)
        return jdate.togregorian()  # یک شیء datetime.date
    except Exception as e:
        print(f"خطا در تبدیل تاریخ: {jdate_str} => {e}")
        return None


def clean_str(value):
    if pd.isna(value) or value is None:
        return ""
    return str(value)
def import_soldiers_from_excel(file_path):
    df = pd.read_excel(file_path, sheet_name=0)
    default_user = MyUser.objects.first() 
    for idx, row in df.iloc[1:].iterrows():
        national_code = str(row.get("کدملی", "")).strip()
        print(f"\n=== ردیف {idx} ===")
        print(f"کد ملی: {national_code}")

        if not national_code:
            print("ردیف بدون کد ملی، نادیده گرفته شد")
            continue

        try:
            with transaction.atomic():
                print("0: شروع پردازش سرباز")

                # آماده‌سازی همه فیلدها قبل از save
                fields = {}

                # کد سازمانی
                org_code = None
                if pd.notna(row.get("کد سازمانی")):
                    org_code, _ = OrganizationalCode.objects.get_or_create(
                        code_number=clean_int(row["کد سازمانی"])
                    )
                    org_code.is_active = True
                    org_code.save()
               #  fields['organizational_code'] = org_code
                print(f"2: کد سازمانی پردازش شد: {org_code}")

                # parent_unit
                parent_unit = None
                if pd.notna(row.get("نام قسمت")):
                    parent_unit, _ = ParentUnit.objects.get_or_create(name=row.get("نام قسمت"))
                fields['current_parent_unit'] = parent_unit
                print(f"3: Parent Unit: {parent_unit}")

                # sub_unit
                sub_unit = None
                if pd.notna(row.get("زیرقسمت")):
                    if parent_unit:
                        sub_unit, _ = SubUnit.objects.get_or_create(
                            name=row.get("زیرقسمت"),
                            parent_unit=parent_unit
                        )
                    else:
                        default_parent, _ = ParentUnit.objects.get_or_create(name="نام قسمت نامشخص")
                        sub_unit, _ = SubUnit.objects.get_or_create(
                            name=row.get("زیرقسمت"),
                            parent_unit=default_parent
                        )
                fields['current_sub_unit'] = sub_unit
                print(f"4: Sub Unit: {sub_unit}")

                # آموزشگاه
                training_center = None
                if pd.notna(row.get("نام آموزشگاه رزم مقدماتی")):
                    training_center, _ = TrainingCenter.objects.get_or_create(
                        name=row.get("نام آموزشگاه رزم مقدماتی")
                    )
                fields['basic_training_center'] = training_center
                print(f"5: Training Center: {training_center}")

                # گروه ناصرین
                naserin_group = None
                if pd.notna(row.get("شماره گروه ناصرین")):
                    naserin_group, created = NaserinGroup.objects.get_or_create(
                        id=clean_int(row.get("شماره گروه ناصرین")),
                        defaults={"manager": default_user}  # اگر نیاز به مقدار پیش‌فرض است
                    )
                fields['naserin_group'] = naserin_group
                print(f"6: Naserin Group: {naserin_group}")

                # دوره عقیدتی
                religious_period = None
                if pd.notna(row.get("دوره عقیدتی")):
                    religious_period, _ = ReligiousPeriod.objects.get_or_create(
                        name=row.get("دوره عقیدتی")
                    )
                fields['ideological_training_period'] = religious_period
                print(f"7: Religious Period: {religious_period}")

                # مقادیر عددی
                fields['number_of_children'] = clean_int(row.get("تعداد اولاد"))
                fields['training_duration'] = clean_int(row.get("مدت آموزش(روز)"))
                fields['service_duration_completed'] = clean_int(row.get("مقدار خدمت انجام شده"))
                fields['total_service_adjustment'] = clean_int(row.get("مجموع (کسری/اضافه خدمت)"))
                fields['fugitive_record'] = clean_int(row.get("سابقه فرار"))
                fields['addiction_record'] = clean_int(row.get("سابقه اعتیاد"))
                fields['number_of_certificates'] = clean_int(row.get("تعداد مدرک"))
                fields['essential_service_duration'] = clean_int(row.get("مدت خدمت ضرورت"))
                fields['degree'] = clean_str(row.get("مدرک"))
                fields['field_of_study'] = clean_str(row.get("رشته تحصیلی"))
                print(f"fields['degree'] = {fields['degree']}")
                print(f"fields['field_of_study'] = {fields['field_of_study']}")
                
                # مقادیر متنی و Boolean
                fields.update({
                    'first_name': clean_str(row.get("نام")),
                    'last_name': clean_str(row.get("نام خانوادگی")),
                    'father_name': clean_str(row.get("نام پدر")),
                    'id_card_code': clean_str(row.get("کد پاسداری")),
                    'birth_date': shamsi_to_gregorian(row.get("تاریخ تولد")),
                    'birth_place': clean_str(row.get("محل تولد")),
                    'issuance_place': clean_str(row.get("محل صدور")),
                    'marital_status': clean_str(row.get("وضعیت تاهل")),
                    'health_status': clean_str(row.get("وضعیت سلامت")),
                    'health_status_description': clean_str(row.get("توضیحات وضعیت سلامت")),
                    'blood_group': clean_str(row.get("گروه خون")),
                    'address': clean_str(row.get("آدرس منزل")),
                    'postal_code': clean_str(row.get("کدپستی")),
                    'phone_home': clean_phone(row.get("منزل")),
                    'phone_mobile': clean_phone(row.get("موبایل")),
                    'phone_virtual': clean_phone(row.get("همراه مجازی")),
                    'phone_parents': clean_phone(row.get("همراه پدر یا مادر")),
                    'rank': clean_str(row.get("درجه")),
                    'is_guard_duty': str(row.get("پاسدار وظیفه", "")).strip() == "بلی",
                    'is_fugitive': str(row.get("فراری", "")).strip() == "بلی",
                    'referral_person': clean_str(row.get("معرف")),
                    'dispatch_date': shamsi_to_gregorian(row.get("تاریخ اعزام")),
                    'service_entry_date': shamsi_to_gregorian(row.get("ورود به یگان")),
                    'service_deduction_type': clean_str(row.get("نوع کسری خدمت")),
                    'service_extension_type': clean_str(row.get("نوع اضافه خدمت")),
                    'service_end_date': shamsi_to_gregorian(row.get("پایان خدمت")),
                    'traffic_status': clean_str(row.get("تردد")),
                    'Is_the_Basij_sufficient': str(row.get("کفایتدار 45 روزه بسیج", "")).strip() == "بلی",
                    'has_driving_license': clean_str(row.get("گواهینامه دارد؟")),
                    'file_shortage': clean_str(row.get("کسری پرونده")),
                    'comments': clean_str(row.get("توضیحات")),
                    'saman_username': clean_str(row.get("نام کاربری ثامن")),
                    'card_chip': clean_str(row.get("تراشه/کارت")),
                    'independent_married': str(row.get("متاهل مستقل", "")) == "بلی",
                    'weekly_or_monthly_presence': clean_str(row.get("حضور هفتگی/ماهانه")),
                    'is_needy': str(row.get("معسرین", "")) == "بلی",
                    'expertise': clean_str(row.get("تخصص")),
                    'is_sunni': str(row.get("اهل تسنن", "")) == "بلی",
                    'is_sayyed': str(row.get("سید", "")) == "بلی",
                    'eligible_for_card_issuance': str(row.get("واجد شرایط صدور کارت پایان خدمت", "")) == "بلی",
                    'card_issuance_status': clean_str(row.get("وضعیت صدور کارت پایان خدمت")),
                    'expired_file_number': clean_str(row.get("شماره پرونده منقضی خدمت")),
                    'skill_5': clean_str(row.get("مهارت 5گانه")),
                    'skill_group': clean_str(row.get("گروه مهارتی")),
                    'skill_certificate': clean_str(row.get("مدرک مهارتی")),
                    'is_certificate': str(row.get("نیاز به مدرک", "")) == "بلی",
                })


                # MultiSelectField برای نوع گواهینامه
                license_value = row.get("نوع گواهینامه", "")
                if pd.isna(license_value) or license_value == "":
                    fields['driving_license_type'] = []
                else:
                    if isinstance(license_value, str):
                        fields['driving_license_type'] = [x.strip() for x in license_value.split(",")]
                    else:
                        fields['driving_license_type'] = [str(license_value)]

                 # گرفتن یا ایجاد سرباز و همزمان مقداردهی اولیه فیلدها
                soldier, created = Soldier.objects.get_or_create(
                  national_code=national_code,
                  defaults=fields  # اینجا دیکشنری fields حاوی تمام فیلدهای دیگر است
                )
                
                soldier.organizational_code = org_code
                org_code.save()


                if not created:
                    # اگر قبلاً وجود داشت و فقط می‌خواهیم به‌روزرسانی کنیم
                   for key, value in fields.items():
                      setattr(soldier, key, value)
                  # ذخیره تغییرات
                soldier.save()
                print(f"8: سرباز با موفقیت ذخیره شد: {soldier}")
        except Exception as e:
            print(f"خطا در ردیف {idx} (کد ملی: {national_code}): {e}")
            import traceback
            traceback.print_exc()
            break


def create_soldiers_excel(soldiers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Soldiers"

    # راست‌به‌چپ کردن کل شیت
    ws.sheet_view.rightToLeft = True

    # هدر جدول
    headers = [
        "ردیف", "کد ملی", "نام", "نام خانوادگی", "نام پدر",
        "وضعیت تاهل", "درجه", "وضعیت سلامت", "گروه مهارتی",
        "مدرک","رشته تحصیلی", "بدهی حقوقی", "آدرس", "موبایل", "شماره کارت پایان خدمت"
    ]
    ws.append(headers)

    # استایل هدر
    header_font = Font(name='B Nazanin', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # داده‌ها + شماره ردیف
    for idx, s in enumerate(soldiers, start=1):
        row = [
            idx,
            s.national_code,
            s.first_name,
            s.last_name,
            s.father_name,
            s.marital_status,
            s.rank,
            s.health_status,
            s.skill_group,
            s.degree,
            s.field_of_study,
            s.file_shortage,
            s.address,
            s.phone_mobile,
            s.card_chip,
        ]
        ws.append(row)

    # چینش متن‌ها راست‌چین و وسط‌چین
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # تنظیم اتوماتیک عرض ستون‌ها
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    return wb


from .constants import RANK_CHOICES
def map_rank_number_to_choice(rank_number: int) -> str | None:
    """
    دریافت عدد 1 تا 12 و بازگشت مقدار متناظر در RANK_CHOICES
    """
    rank_map = {i+1: choice[0] for i, choice in enumerate(RANK_CHOICES)}
    return rank_map.get(rank_number)
